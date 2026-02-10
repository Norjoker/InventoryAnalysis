import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

from inventory_aggregator import aggregate_inventory_by_sn, write_serial_history_workbook


class TestInventoryAggregator(unittest.TestCase):
    @staticmethod
    def _write_snapshot(path: Path, rows: list[list[object]]) -> None:
        frame = pd.DataFrame(rows, columns=["A", "B", "SN", "D", "E", "F", "G"])
        frame.to_excel(path, index=False, engine="openpyxl")

    def test_aggregate_inventory_tracks_first_and_last_rows(self) -> None:
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            snap_1 = temp_path / "2024-01-01_Raw_Data.xlsx"
            snap_2 = temp_path / "2024-02-01_Raw_Data.xlsx"

            self._write_snapshot(
                snap_1,
                [
                    ["A1", "B1", "SN1", "D1", "E1", "F1", "G1"],
                    ["A2", "B2", "SN2", "D2", "E2", "F2", "G2"],
                    ["A3", "B3", "", "D3", "E3", "F3", "G3"],
                ],
            )
            self._write_snapshot(
                snap_2,
                [
                    ["A1-NEW", "B1-NEW", "SN1", "D1-NEW", "E1-NEW", "F1-NEW", "G1-NEW"],
                    ["A4", "B4", "SN3", "D4", "E4", "F4", "G4"],
                ],
            )

            result = aggregate_inventory_by_sn(
                [
                    (snap_1, date(2024, 1, 1)),
                    (snap_2, date(2024, 2, 1)),
                ]
            )

            sn1 = result[result["sn"] == "SN1"].iloc[0]
            self.assertEqual(sn1["first_seen"], date(2024, 1, 1))
            self.assertEqual(sn1["last_seen"], date(2024, 2, 1))
            self.assertEqual(sn1["first_col_a"], "A1")
            self.assertEqual(sn1["last_col_a"], "A1-NEW")

            sn2 = result[result["sn"] == "SN2"].iloc[0]
            self.assertEqual(sn2["first_seen"], date(2024, 1, 1))
            self.assertEqual(sn2["last_seen"], date(2024, 1, 1))

            sn3 = result[result["sn"] == "SN3"].iloc[0]
            self.assertEqual(sn3["first_seen"], date(2024, 2, 1))
            self.assertEqual(sn3["last_seen"], date(2024, 2, 1))

    def test_aggregate_inventory_validates_column_c_header(self) -> None:
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            bad_snapshot = temp_path / "2024-03-01_Raw_Data.xlsx"
            bad_frame = pd.DataFrame(
                [[1, 2, 3, 4, 5, 6, 7]],
                columns=["A", "B", "SERIAL", "D", "E", "F", "G"],
            )
            bad_frame.to_excel(bad_snapshot, index=False, engine="openpyxl")

            with self.assertRaisesRegex(ValueError, "Column C header must be 'SN'"):
                aggregate_inventory_by_sn([(bad_snapshot, date(2024, 3, 1))])

    def test_write_serial_history_workbook_exports_excel_dates_and_formatting(self) -> None:
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            snap_1 = temp_path / "2024-01-01_Raw_Data.xlsx"
            snap_2 = temp_path / "2024-02-01_Raw_Data.xlsx"
            output = temp_path / "compiled_inventory_serial_history.xlsx"

            self._write_snapshot(
                snap_1,
                [["A1", "B1", "SN1", "D1", "E1", "F1", "G1"]],
            )
            self._write_snapshot(
                snap_2,
                [["A1-NEW", "B1-NEW", "SN1", "D1-NEW", "E1-NEW", "F1-NEW", "G1-NEW"]],
            )

            write_serial_history_workbook(
                [
                    (snap_1, date(2024, 1, 1)),
                    (snap_2, date(2024, 2, 1)),
                ],
                output_path=output,
                include_run_log=True,
            )

            workbook = load_workbook(output)
            self.assertIn("Serial_History", workbook.sheetnames)
            self.assertIn("Run_Log", workbook.sheetnames)

            serial_ws = workbook["Serial_History"]
            run_log_ws = workbook["Run_Log"]

            self.assertEqual(serial_ws.freeze_panes, "A2")
            self.assertIsNotNone(serial_ws.auto_filter.ref)
            self.assertEqual(run_log_ws.freeze_panes, "A2")
            self.assertIsNotNone(run_log_ws.auto_filter.ref)

            first_seen_cell = serial_ws.cell(row=2, column=2)
            last_seen_cell = serial_ws.cell(row=2, column=3)
            run_log_date_cell = run_log_ws.cell(row=2, column=2)

            self.assertEqual(first_seen_cell.number_format, "yyyy-mm-dd")
            self.assertEqual(last_seen_cell.number_format, "yyyy-mm-dd")
            self.assertEqual(run_log_date_cell.number_format, "yyyy-mm-dd")

            self.assertEqual(first_seen_cell.value.date(), date(2024, 1, 1))
            self.assertEqual(last_seen_cell.value.date(), date(2024, 2, 1))
            self.assertEqual(run_log_date_cell.value.date(), date(2024, 1, 1))


if __name__ == "__main__":
    unittest.main()
